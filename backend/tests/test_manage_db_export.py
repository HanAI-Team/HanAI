import csv
import io
from datetime import datetime, timezone
from uuid import UUID

import pytest_asyncio
from openpyxl import load_workbook

from app.auth.service import create_access_token
from app.core.models import Doctor, Hospital, Patient


@pytest_asyncio.fixture
async def associate_doctor(db, approved_doctor):
    owner, _ = approved_doctor
    doctor = Doctor(
        hospital_id=owner.hospital_id,
        name="부원장",
        license_number="11112222",
        password_hash="hashed_password",
        role="associate",
        is_approved=True,
        approved_at=datetime.now(timezone.utc),
    )
    db.add(doctor)
    await db.commit()
    await db.refresh(doctor)
    token = create_access_token(UUID(str(doctor.id)), UUID(str(doctor.hospital_id)), "associate")
    return doctor, {"Authorization": f"Bearer {token}"}


def _csv_rows(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig")
    return list(csv.reader(io.StringIO(text)))


async def test_db_export_rejects_non_owner(client, associate_doctor):
    _, headers = associate_doctor
    resp = await client.get(
        "/api/manage/db-export",
        params={"table": "patients", "reason": "테스트"},
        headers=headers,
    )
    assert resp.status_code == 403


async def test_db_export_rejects_unknown_table(client, approved_doctor):
    _, headers = approved_doctor
    resp = await client.get(
        "/api/manage/db-export",
        params={"table": "not_a_real_table", "reason": "테스트"},
        headers=headers,
    )
    assert resp.status_code == 400


async def test_db_export_requires_reason(client, approved_doctor):
    _, headers = approved_doctor
    resp = await client.get("/api/manage/db-export", params={"table": "patients"}, headers=headers)
    assert resp.status_code == 422


async def test_db_export_patients_masks_rrn_and_scopes_by_hospital(client, approved_doctor, db):
    doctor, headers = approved_doctor
    db.add(Patient(hospital_id=doctor.hospital_id, name="내병원환자", rrn="9001011234567"))

    other_hospital = Hospital(name="다른병원")
    db.add(other_hospital)
    await db.flush()
    db.add(Patient(hospital_id=other_hospital.id, name="다른병원환자", rrn="9001011234568"))
    await db.commit()

    resp = await client.get(
        "/api/manage/db-export",
        params={"table": "patients", "reason": "HIRA 기능검사"},
        headers=headers,
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")
    rows = _csv_rows(resp.content)
    header = rows[0]
    body = rows[1:]
    assert len(body) == 1
    name_idx = header.index("name")
    rrn_idx = header.index("rrn")
    assert body[0][name_idx] == "내병원환자"
    assert body[0][rrn_idx] == "***-*******"


async def test_db_export_xlsx_format(client, approved_doctor, db):
    doctor, headers = approved_doctor
    db.add(Patient(hospital_id=doctor.hospital_id, name="엑셀환자"))
    await db.commit()

    resp = await client.get(
        "/api/manage/db-export",
        params={"table": "patients", "format": "xlsx", "reason": "HIRA 기능검사"},
        headers=headers,
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "id"
    values = [cell for row in rows[1:] for cell in row]
    assert "엑셀환자" in values


async def test_db_export_global_master_table_not_hospital_filtered(client, approved_doctor, db):
    from app.core.models import FeeMaster

    doctor, headers = approved_doctor
    db.add(FeeMaster(code="ZZ900", name="테스트수가", category="침술", unit_price=1000))
    await db.commit()

    resp = await client.get(
        "/api/manage/db-export",
        params={"table": "fee_master", "reason": "HIRA 기능검사"},
        headers=headers,
    )
    assert resp.status_code == 200
    rows = _csv_rows(resp.content)
    codes = [r[rows[0].index("code")] for r in rows[1:]]
    assert "ZZ900" in codes


async def test_db_export_logs_download(client, approved_doctor, db):
    from sqlalchemy import select

    from app.core.models import DataDownloadLog

    doctor, headers = approved_doctor
    resp = await client.get(
        "/api/manage/db-export",
        params={"table": "medical_records", "reason": "기능검사 시연"},
        headers=headers,
    )
    assert resp.status_code == 200

    result = await db.execute(
        select(DataDownloadLog).where(DataDownloadLog.hospital_id == doctor.hospital_id)
    )
    log = result.scalar_one()
    assert log.download_type == "db_export:medical_records"
    assert log.reason == "기능검사 시연"


async def test_db_export_every_whitelisted_table_returns_200(client, approved_doctor):
    from app.manage.router import TABLE_WHITELIST

    _, headers = approved_doctor
    for table in TABLE_WHITELIST:
        resp = await client.get(
            "/api/manage/db-export",
            params={"table": table, "reason": "화이트리스트 점검"},
            headers=headers,
        )
        assert resp.status_code == 200, f"{table} failed: {resp.text}"
